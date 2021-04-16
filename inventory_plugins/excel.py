from ansible.plugins.inventory import BaseInventoryPlugin

import openpyxl

ANSIBLE_METADATA = {
    'metadata_version': '0.9.0',
    'status': ['preview'],
    'supported_by': 'community'
}

DOCUMENTATION = r'''
---
module: excel
plugin_type: inventory
short_description: Excel Inventory

options:
    excel_file:
        description: Excel file to use
        default: inventory.xlsx
        required: false
        type: str
    hosts_groups:
        description: add inventory to this group
        default: _all
        required: false
        type: str
    layout:
        description: Excel layout 
        default: sheet_header
        required: false
        type: str
author:
    - Greg I/O
'''

class InventoryModule(BaseInventoryPlugin):
    """Excel Inventory."""

    NAME = 'excel'

    def verify_file(self, path):
        """Verify that the source file can be processed correctly.

        Parameters:
            path:AnyStr The path to the file that needs to be verified

        Returns:
            bool True if the file is valid, else False
        """
        valid = False
        if super(InventoryModule, self).verify_file(path):
            if path.endswith(('excel.yaml')):
                valid = True

        return valid

    def read_xls_dict(self, input_file):
        "Read the XLS file and return as Ansible facts"
        spreadsheet = {}
        try:
            wb = openpyxl.load_workbook(input_file, data_only=True)
            for sheet in wb.get_sheet_names():
                ansible_sheet_name = sheet.lower().replace(' ','_')
                spreadsheet[ansible_sheet_name] = []
                current_sheet = wb.get_sheet_by_name(sheet)
                dict_keys = []
                for c in range(1,current_sheet.max_column + 1):
                    dict_keys.append(current_sheet.cell(row=1,column=c).value)
                for r in range (2,current_sheet.max_row + 1):
                    temp_dict = {}
                    for c in range(1,current_sheet.max_column + 1):
                        temp_dict[dict_keys[c-1].lower().replace(' ','_')] = current_sheet.cell(row=r,column=c).value
                    spreadsheet[ansible_sheet_name].append(temp_dict)
        except IOError:
            return (1, "IOError on input file:%s" % input_file)

        return (0,spreadsheet)

    def sheet_header(self, inventory):
        """ 

        """
        excel_file = self.get_option('excel_file')
        hosts_groups = self.get_option('hosts_groups').split(",")

        for i in hosts_groups:
            if i.find('_') == 0:
                self.inventory.add_group(i.replace('_','', 1))
            else:
                self.inventory.add_host(i)

        ret, excel = self.read_xls_dict(excel_file)

        for key in excel.keys():
                [self.inventory.set_variable(i.replace('_','',1), key, excel[key]) for i in hosts_groups]
                

    def parse(self, inventory, loader, path, cache):
        """Parse and populate the inventory with data about hosts.

        Parameters:
            inventory The inventory to populate

        We ignore the other parameters in the future signature, as we will
        not use them.

        Returns:
            None
        """
        super(InventoryModule, self).parse(inventory, loader, path)
        # this method will parse 'common format' inventory sources and
        # update any options declared in DOCUMENTATION as needed
        config = self._read_config_data(path)

        layout = self.get_option('layout')

        class_method = getattr(self, layout)

        class_method(inventory)
